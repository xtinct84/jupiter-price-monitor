"""
export_backtest_session.py
══════════════════════════════════════════════════════════════════════
Exports signal backtest data covering only the period AFTER the most
recent cold-start warm-up completed.

Warm-up end is derived automatically from the DB:
  session_start = earliest quote timestamp within QUOTE_MAX_AGE_HOURS
  warmup_end    = session_start + WARMUP_ITERATIONS × 30s

Only signals logged after warmup_end are included — cold-start false
positives (like the multi-signal cascade on Iteration 1) are excluded.

Output: Excel workbook with 4 sheets
  1. Session Summary   — warm-up timeline + signal counts + score dist.
  2. Signals           — every post-warmup signal with full detail
  3. Backtest          — signals with forward price outcomes (5/15/30min)
  4. Pair Performance  — per-pair accuracy and profit metrics

Run from your project root (same folder as price_history/):
  python export_backtest_session.py              # auto-detect session (last 8h)
  python export_backtest_session.py --hours 12  # look back 12h for session anchor

If you run the export hours after a monitoring session ended, use --hours to
extend the lookback window so the session anchor quotes are still found.

Output file: price_history/backtest_session_<timestamp>.xlsx
══════════════════════════════════════════════════════════════════════
"""

import argparse
import sqlite3
import pandas as pd
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
import sys

# ── Match constants from arbitrage_detector.py ───────────────────────────────
QUOTE_MAX_AGE_HOURS  = 2        # detector uses 2h — exporter default extends this
WARMUP_ITERATIONS    = 20
ITERATION_INTERVAL_S = 30   # seconds between monitor cycles
WARMUP_DURATION_S    = WARMUP_ITERATIONS * ITERATION_INTERVAL_S  # 600s = 10min

# ── DB path ───────────────────────────────────────────────────────────────────
DB_PATH = Path("price_history/jupiter_monitor.db")

# ── Style constants ───────────────────────────────────────────────────────────
FONT_NAME   = "Arial"
COL_HEADER  = "1F3864"   # dark navy
COL_WARMUP  = "FFF2CC"   # soft yellow
COL_LIVE    = "E2EFDA"   # soft green
COL_EXECUTE = "C6EFCE"   # bright green
COL_ANALYSIS= "FFEB9C"   # amber
COL_REJECT  = "FFC7CE"   # red
COL_ALT_ROW = "F5F5F5"   # light grey alternate row


def thin_border():
    s = Side(border_style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)


def header_font(bold=True):
    return Font(name=FONT_NAME, bold=bold, color="FFFFFF", size=10)


def body_font(bold=False):
    return Font(name=FONT_NAME, bold=bold, size=10)


def hfill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)


def set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def style_header_row(ws, row_num, n_cols, fill_hex=COL_HEADER):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.font      = header_font()
        cell.fill      = hfill(fill_hex)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border()


def write_kv(ws, row, label, value, label_bold=True, value_color=None):
    lc = ws.cell(row=row, column=1, value=label)
    lc.font = Font(name=FONT_NAME, bold=label_bold, size=10)
    vc = ws.cell(row=row, column=2, value=value)
    vc.font = Font(name=FONT_NAME, size=10,
                   color=value_color if value_color else "000000")
    vc.alignment = Alignment(horizontal="left")


# ── Database helpers ──────────────────────────────────────────────────────────

def get_conn():
    if not DB_PATH.exists():
        print(f"❌  Database not found: {DB_PATH}")
        print("    Run from your project root alongside price_history/")
        sys.exit(1)
    return sqlite3.connect(DB_PATH)


def derive_session_window(conn, lookback_hours: int = 8):
    """
    Find the most recent monitoring session start and compute warm-up end.

    Strategy:
      1. Look for the most recent gap in quotes >= 5 minutes within lookback_hours.
         A gap means a restart — the quote immediately after the gap is the
         session start.
      2. If no gap found, the earliest quote in the window is the session start
         (continuous session).
      3. If no quotes found at all within lookback_hours, extends search to the
         most recent quote in the entire DB as a last resort.

    Args:
        conn:           SQLite connection
        lookback_hours: How far back to search for session anchor quotes.
                        Default 8h. Pass --hours N on CLI if session is older.

    Returns:
        (session_start: datetime, warmup_end: datetime, quote_count_in_session: int)
    """
    cursor = conn.cursor()
    cutoff = datetime.now() - timedelta(hours=lookback_hours)
    cutoff_str = cutoff.strftime("%Y-%m-%d %H:%M:%S")

    # Fetch all distinct quote timestamps in the lookback window, ordered ASC
    cursor.execute("""
        SELECT DISTINCT timestamp FROM quotes
        WHERE timestamp >= ?
        ORDER BY timestamp ASC
    """, (cutoff_str,))
    rows = cursor.fetchall()

    if not rows:
        # No quotes in lookback window — try to find any quotes in DB at all
        cursor.execute("SELECT MAX(timestamp) FROM quotes")
        last = cursor.fetchone()[0]
        if last:
            # Find earliest quote in a 3h window around the last known quote
            last_dt  = datetime.strptime(last, "%Y-%m-%d %H:%M:%S")
            fallback = (last_dt - timedelta(hours=3)).strftime("%Y-%m-%d %H:%M:%S")
            cursor.execute("""
                SELECT MIN(timestamp), COUNT(DISTINCT timestamp) FROM quotes
                WHERE timestamp >= ?
            """, (fallback,))
            fb_row = cursor.fetchone()
            session_start = datetime.strptime(fb_row[0], "%Y-%m-%d %H:%M:%S")
            q_count       = fb_row[1] or 0
            print(f"  ⚠️  No quotes in last {lookback_hours}h — "
                  f"falling back to last DB session (~{last_dt.strftime('%H:%M')})")
        else:
            # DB is completely empty
            session_start = datetime.now() - timedelta(minutes=10)
            q_count = 0
        warmup_end = session_start + timedelta(seconds=WARMUP_DURATION_S)
        return session_start, warmup_end, q_count

    timestamps = [datetime.strptime(r[0], "%Y-%m-%d %H:%M:%S") for r in rows]

    # Find the most recent restart gap (>= 5 min silence between consecutive quotes)
    GAP_THRESHOLD_S = 300  # 5 minutes = likely a restart
    session_start   = timestamps[0]   # default: start of lookback window

    for i in range(len(timestamps) - 1, 0, -1):
        gap = (timestamps[i] - timestamps[i - 1]).total_seconds()
        if gap >= GAP_THRESHOLD_S:
            session_start = timestamps[i]  # first quote after the restart gap
            break

    warmup_end = session_start + timedelta(seconds=WARMUP_DURATION_S)

    # Count quotes from session_start onward
    cursor.execute("""
        SELECT COUNT(DISTINCT timestamp) FROM quotes
        WHERE timestamp >= ?
    """, (session_start.strftime("%Y-%m-%d %H:%M:%S"),))
    q_count = cursor.fetchone()[0] or 0

    return session_start, warmup_end, q_count


def load_signals(conn, warmup_end: datetime) -> pd.DataFrame:
    """Load all signals logged after the warm-up ended."""
    warmup_end_str = warmup_end.strftime("%Y-%m-%d %H:%M:%S")
    df = pd.read_sql_query("""
        SELECT
            id,
            timestamp,
            signal_type,
            pair,
            description,
            estimated_profit_pct,
            weighted_score,
            execute_candidate,
            condition_breakdown,
            resolved
        FROM signals
        WHERE timestamp >= ?
        ORDER BY timestamp ASC
    """, conn, params=(warmup_end_str,))
    if not df.empty:
        df["timestamp"] = pd.to_datetime(df["timestamp"])
        df["execute_candidate"] = df["execute_candidate"].astype(bool)
    return df


def load_warmup_signals(conn, session_start: datetime, warmup_end: datetime) -> pd.DataFrame:
    """Load signals that fired DURING warm-up (for reference / exclusion audit)."""
    df = pd.read_sql_query("""
        SELECT timestamp, signal_type, pair, estimated_profit_pct, weighted_score
        FROM signals
        WHERE timestamp >= ? AND timestamp < ?
        ORDER BY timestamp ASC
    """, conn, params=(
        session_start.strftime("%Y-%m-%d %H:%M:%S"),
        warmup_end.strftime("%Y-%m-%d %H:%M:%S")
    ))
    if not df.empty:
        df["timestamp"] = pd.to_datetime(df["timestamp"])
    return df


def load_prices(conn) -> pd.DataFrame:
    """Load all price records for forward-outcome calculation."""
    df = pd.read_sql_query("""
        SELECT timestamp, symbol, price_usd
        FROM prices
        ORDER BY timestamp ASC
    """, conn)
    if not df.empty:
        df["timestamp"] = pd.to_datetime(df["timestamp"])
    return df


def get_forward_price(prices_df, symbol, signal_ts, minutes_ahead):
    """Return price of symbol N minutes after signal_ts, or None."""
    target = signal_ts + timedelta(minutes=minutes_ahead)
    window = prices_df[
        (prices_df["symbol"] == symbol) &
        (prices_df["timestamp"] >= target)
    ]
    if window.empty:
        return None
    return window.iloc[0]["price_usd"]


def build_backtest(signals_df: pd.DataFrame, prices_df: pd.DataFrame) -> pd.DataFrame:
    """
    Attach forward price outcomes to each signal.
    For rate_divergence / impact_anomaly: base symbol price (first token in pair).
    For triangular: base symbol is first token in path.
    """
    rows = []
    for _, sig in signals_df.iterrows():
        pair = sig["pair"]
        # Extract base symbol (e.g. SOL from SOL/USDC or SOL→USDC→USDT→SOL)
        base = pair.split("/")[0].split("→")[0].strip()
        ts   = sig["timestamp"]

        price_at_signal = get_forward_price(prices_df, base, ts, 0)
        price_5m        = get_forward_price(prices_df, base, ts, 5)
        price_15m       = get_forward_price(prices_df, base, ts, 15)
        price_30m       = get_forward_price(prices_df, base, ts, 30)

        def pct_move(p_future):
            if price_at_signal and p_future and price_at_signal > 0:
                return round((p_future - price_at_signal) / price_at_signal * 100, 4)
            return None

        rows.append({
            "timestamp":            ts,
            "signal_type":          sig["signal_type"],
            "pair":                 pair,
            "base_symbol":          base,
            "estimated_profit_pct": sig["estimated_profit_pct"],
            "weighted_score":       sig["weighted_score"],
            "execute_candidate":    sig["execute_candidate"],
            "price_at_signal":      round(price_at_signal, 6) if price_at_signal else None,
            "price_5m":             round(price_5m,  6) if price_5m  else None,
            "price_15m":            round(price_15m, 6) if price_15m else None,
            "price_30m":            round(price_30m, 6) if price_30m else None,
            "move_5m_pct":          pct_move(price_5m),
            "move_15m_pct":         pct_move(price_15m),
            "move_30m_pct":         pct_move(price_30m),
            "direction_5m":         ("UP" if pct_move(price_5m) and pct_move(price_5m) > 0
                                     else "DOWN" if pct_move(price_5m) is not None else None),
            "direction_15m":        ("UP" if pct_move(price_15m) and pct_move(price_15m) > 0
                                     else "DOWN" if pct_move(price_15m) is not None else None),
            "direction_30m":        ("UP" if pct_move(price_30m) and pct_move(price_30m) > 0
                                     else "DOWN" if pct_move(price_30m) is not None else None),
        })
    return pd.DataFrame(rows)


def build_pair_performance(backtest_df: pd.DataFrame) -> pd.DataFrame:
    """Aggregate per-pair accuracy and profit metrics."""
    if backtest_df.empty:
        return pd.DataFrame()

    rows = []
    for pair, grp in backtest_df.groupby("pair"):
        n = len(grp)
        ex = grp["execute_candidate"].sum()

        def acc(col):
            valid = grp[col].dropna()
            if valid.empty:
                return None
            return round((valid > 0).sum() / len(valid) * 100, 1)

        def avg_move(col):
            valid = grp[col].dropna()
            return round(valid.mean(), 4) if not valid.empty else None

        rows.append({
            "pair":                   pair,
            "signal_count":           n,
            "execute_candidates":     int(ex),
            "avg_estimated_profit":   round(grp["estimated_profit_pct"].mean(), 4),
            "avg_score":              round(grp["weighted_score"].mean(), 1),
            "accuracy_5m_pct":        acc("move_5m_pct"),
            "accuracy_15m_pct":       acc("move_15m_pct"),
            "accuracy_30m_pct":       acc("move_30m_pct"),
            "avg_move_5m_pct":        avg_move("move_5m_pct"),
            "avg_move_15m_pct":       avg_move("move_15m_pct"),
            "avg_move_30m_pct":       avg_move("move_30m_pct"),
        })
    df = pd.DataFrame(rows).sort_values("signal_count", ascending=False)
    return df


# ── Sheet writers ─────────────────────────────────────────────────────────────

def write_summary_sheet(wb, session_start, warmup_end, signals_df,
                         warmup_signals_df, fresh_quote_count):
    ws = wb.active
    ws.title = "Session Summary"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A": 32, "B": 28, "C": 18, "D": 18})

    row = 1
    # Title banner
    ws.merge_cells(f"A{row}:D{row}")
    tc = ws.cell(row=row, column=1,
                 value="📊  JUPITER ARBITRAGE — POST WARM-UP BACKTEST")
    tc.font      = Font(name=FONT_NAME, bold=True, size=13, color="FFFFFF")
    tc.fill      = hfill(COL_HEADER)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28
    row += 1

    ws.merge_cells(f"A{row}:D{row}")
    sc = ws.cell(row=row, column=1,
                 value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    sc.font      = Font(name=FONT_NAME, italic=True, size=9, color="666666")
    sc.fill      = hfill("F0F4FF")
    sc.alignment = Alignment(horizontal="center")
    row += 2

    # Session timeline section
    def section_header(label, r):
        ws.merge_cells(f"A{r}:D{r}")
        c = ws.cell(row=r, column=1, value=f"  {label}")
        c.font  = Font(name=FONT_NAME, bold=True, size=10, color="FFFFFF")
        c.fill  = hfill("2E4057")
        c.alignment = Alignment(vertical="center")
        ws.row_dimensions[r].height = 20
        return r + 1

    row = section_header("SESSION TIMELINE", row)

    timeline = [
        ("Session start (first fresh quote)",
         session_start.strftime("%Y-%m-%d %H:%M:%S"), COL_WARMUP),
        (f"Warm-up duration ({WARMUP_ITERATIONS} iterations × {ITERATION_INTERVAL_S}s)",
         f"{WARMUP_DURATION_S}s  ({WARMUP_DURATION_S // 60} min)", COL_WARMUP),
        ("Warm-up end (signals enabled from)",
         warmup_end.strftime("%Y-%m-%d %H:%M:%S"), COL_LIVE),
        ("Signals suppressed during warm-up",
         len(warmup_signals_df), None),
        ("Quotes counted in session window",
         fresh_quote_count, None),
    ]
    for label, value, fill in timeline:
        lc = ws.cell(row=row, column=1, value=label)
        vc = ws.cell(row=row, column=2, value=value)
        lc.font = body_font()
        vc.font = body_font(bold=True)
        if fill:
            lc.fill = hfill(fill)
            vc.fill = hfill(fill)
        lc.border = vc.border = thin_border()
        vc.alignment = Alignment(horizontal="left")
        row += 1

    row += 1
    row = section_header("SIGNAL COUNTS (POST WARM-UP)", row)

    if signals_df.empty:
        ws.cell(row=row, column=1,
                value="⚠️  No post-warm-up signals found in database.").font = \
            Font(name=FONT_NAME, italic=True, color="FF0000")
        row += 1
    else:
        counts = [
            ("Total signals",        len(signals_df)),
            ("Execute candidates",   int(signals_df["execute_candidate"].sum())),
            ("Analysis only",        int((~signals_df["execute_candidate"]).sum())),
        ]
        for label, value in counts:
            lc = ws.cell(row=row, column=1, value=label)
            vc = ws.cell(row=row, column=2, value=value)
            lc.font = body_font()
            vc.font = body_font(bold=True)
            lc.border = vc.border = thin_border()
            row += 1

        row += 1
        row = section_header("SIGNALS BY TYPE", row)
        for stype, grp in signals_df.groupby("signal_type"):
            lc = ws.cell(row=row, column=1, value=stype.replace("_", " ").title())
            vc = ws.cell(row=row, column=2, value=len(grp))
            ec = ws.cell(row=row, column=3,
                         value=f"Avg score: {grp['weighted_score'].mean():.1f}")
            pc = ws.cell(row=row, column=4,
                         value=f"Avg profit: {grp['estimated_profit_pct'].mean():.3f}%")
            for c in (lc, vc, ec, pc):
                c.font   = body_font()
                c.border = thin_border()
            row += 1

        row += 1
        row = section_header("SCORE DISTRIBUTION (POST WARM-UP)", row)
        bins = [(55, 74, "Analysis gate (55–74)"),
                (75, 84, "Execute gate (75–84)"),
                (85, 100, "High confidence (85–100)")]
        for lo, hi, label in bins:
            count = int(((signals_df["weighted_score"] >= lo) &
                         (signals_df["weighted_score"] <= hi)).sum())
            lc = ws.cell(row=row, column=1, value=label)
            vc = ws.cell(row=row, column=2, value=count)
            lc.font = body_font()
            vc.font = body_font(bold=True)
            lc.border = vc.border = thin_border()
            row += 1

    if not warmup_signals_df.empty:
        row += 1
        row = section_header("WARM-UP PERIOD — EXCLUDED SIGNALS (for audit)", row)
        ws.cell(row=row, column=1,
                value="These signals fired during warm-up and were suppressed.").font = \
            Font(name=FONT_NAME, italic=True, size=9, color="888888")
        row += 1
        headers = ["Timestamp", "Type", "Pair", "Est. Profit %", "Score"]
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font  = Font(name=FONT_NAME, bold=True, size=9, color="FFFFFF")
            c.fill  = hfill("888888")
            c.border = thin_border()
        row += 1
        for _, sig in warmup_signals_df.iterrows():
            vals = [str(sig["timestamp"])[:19], sig["signal_type"],
                    sig["pair"], sig["estimated_profit_pct"], sig["weighted_score"]]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=v)
                c.font   = Font(name=FONT_NAME, size=9, color="888888")
                c.fill   = hfill("F9F9F9")
                c.border = thin_border()
            row += 1


def write_signals_sheet(wb, signals_df):
    ws = wb.create_sheet("Signals")
    ws.sheet_view.showGridLines = False

    if signals_df.empty:
        ws.cell(row=1, column=1, value="No post-warm-up signals.")
        return

    headers = [
        "Timestamp", "Type", "Pair", "Est. Profit %",
        "Score", "Execute Candidate", "Description"
    ]
    col_widths = {"A": 20, "B": 20, "C": 22, "D": 15,
                  "E": 10, "F": 18, "G": 60}
    set_col_widths(ws, col_widths)

    ws.row_dimensions[1].height = 22
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font      = header_font()
        c.fill      = hfill(COL_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin_border()
    ws.freeze_panes = "A2"

    for ri, (_, sig) in enumerate(signals_df.iterrows(), 2):
        is_exec = sig["execute_candidate"]
        row_fill = hfill(COL_EXECUTE if is_exec else
                         (COL_ALT_ROW if ri % 2 == 0 else "FFFFFF"))
        vals = [
            str(sig["timestamp"])[:19],
            sig["signal_type"].replace("_", " ").title(),
            sig["pair"],
            round(sig["estimated_profit_pct"], 4),
            int(sig["weighted_score"]),
            "✅ EXECUTE" if is_exec else "📊 ANALYSIS",
            str(sig.get("description", ""))[:200],
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font   = body_font(bold=is_exec)
            c.fill   = row_fill
            c.border = thin_border()
            c.alignment = Alignment(horizontal="left", vertical="center",
                                    wrap_text=(ci == 7))
        # Color profit cell
        ws.cell(row=ri, column=4).number_format = "0.0000%"


def write_backtest_sheet(wb, backtest_df):
    ws = wb.create_sheet("Backtest")
    ws.sheet_view.showGridLines = False

    if backtest_df.empty:
        ws.cell(row=1, column=1, value="No backtest data available.")
        return

    headers = [
        "Timestamp", "Type", "Pair", "Base Symbol",
        "Est. Profit %", "Score", "Execute",
        "Price @ Signal", "Price +5m", "Price +15m", "Price +30m",
        "Move +5m %", "Move +15m %", "Move +30m %",
        "Direction 5m", "Direction 15m", "Direction 30m",
    ]
    widths = {
        "A": 20, "B": 20, "C": 22, "D": 13,
        "E": 13, "F": 8,  "G": 10,
        "H": 14, "I": 14, "J": 14, "K": 14,
        "L": 12, "M": 12, "N": 12,
        "O": 13, "P": 13, "Q": 13,
    }
    set_col_widths(ws, widths)

    ws.row_dimensions[1].height = 22
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font      = header_font()
        c.fill      = hfill(COL_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border    = thin_border()
    ws.freeze_panes = "A2"

    pct_cols = {12, 13, 14}  # move % columns (1-indexed)

    for ri, (_, row) in enumerate(backtest_df.iterrows(), 2):
        is_exec  = row.get("execute_candidate", False)
        alt_fill = hfill(COL_ALT_ROW if ri % 2 == 0 else "FFFFFF")

        vals = [
            str(row["timestamp"])[:19],
            row["signal_type"].replace("_", " ").title(),
            row["pair"],
            row["base_symbol"],
            round(row["estimated_profit_pct"], 4),
            int(row["weighted_score"]),
            "✅" if is_exec else "📊",
            row.get("price_at_signal"),
            row.get("price_5m"),
            row.get("price_15m"),
            row.get("price_30m"),
            row.get("move_5m_pct"),
            row.get("move_15m_pct"),
            row.get("move_30m_pct"),
            row.get("direction_5m"),
            row.get("direction_15m"),
            row.get("direction_30m"),
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.border = thin_border()
            c.alignment = Alignment(horizontal="center" if ci in {7, 15, 16, 17}
                                    else "left", vertical="center")

            # Direction coloring
            if ci in {15, 16, 17} and v:
                c.font = Font(name=FONT_NAME, bold=True, size=10,
                              color="375623" if v == "UP" else "9C0006")
                c.fill = hfill("C6EFCE" if v == "UP" else "FFC7CE")
            else:
                c.font = body_font(bold=is_exec)
                c.fill = hfill(COL_EXECUTE) if is_exec else alt_fill

            # Format price columns to 6dp
            if ci in {8, 9, 10, 11} and v is not None:
                c.number_format = "#,##0.000000"
            if ci in pct_cols and v is not None:
                c.number_format = "0.0000"

    # Conditional formatting on move % columns (green-white-red)
    last_row = len(backtest_df) + 1
    for col_idx in [12, 13, 14]:
        col_letter = get_column_letter(col_idx)
        cell_range = f"{col_letter}2:{col_letter}{last_row}"
        ws.conditional_formatting.add(cell_range, ColorScaleRule(
            start_type="min", start_color="FFC7CE",
            mid_type="num", mid_value=0, mid_color="FFFFFF",
            end_type="max", end_color="C6EFCE",
        ))


def write_pair_performance_sheet(wb, pair_df, backtest_df):
    ws = wb.create_sheet("Pair Performance")
    ws.sheet_view.showGridLines = False

    if pair_df.empty:
        ws.cell(row=1, column=1, value="No pair performance data available.")
        return

    # ── Top section: per-pair table ──
    headers = [
        "Pair", "Signals", "Execute Candidates",
        "Avg Est. Profit %", "Avg Score",
        "Accuracy @5m %", "Accuracy @15m %", "Accuracy @30m %",
        "Avg Move @5m %", "Avg Move @15m %", "Avg Move @30m %",
    ]
    widths = {
        "A": 22, "B": 11, "C": 20, "D": 18, "E": 12,
        "F": 16, "G": 16, "H": 16,
        "I": 16, "J": 16, "K": 16,
    }
    set_col_widths(ws, widths)

    ws.row_dimensions[1].height = 22
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font      = header_font()
        c.fill      = hfill(COL_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border    = thin_border()
    ws.freeze_panes = "A2"

    for ri, (_, row) in enumerate(pair_df.iterrows(), 2):
        fill = hfill(COL_ALT_ROW if ri % 2 == 0 else "FFFFFF")
        vals = [
            row["pair"],
            int(row["signal_count"]),
            int(row["execute_candidates"]),
            row["avg_estimated_profit"],
            row["avg_score"],
            row.get("accuracy_5m_pct"),
            row.get("accuracy_15m_pct"),
            row.get("accuracy_30m_pct"),
            row.get("avg_move_5m_pct"),
            row.get("avg_move_15m_pct"),
            row.get("avg_move_30m_pct"),
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font   = body_font()
            c.fill   = fill
            c.border = thin_border()
            c.alignment = Alignment(horizontal="left" if ci == 1 else "center",
                                    vertical="center")
            if v is not None and ci in {6, 7, 8}:
                c.number_format = "0.0"
            if v is not None and ci in {4, 9, 10, 11}:
                c.number_format = "0.0000"

    # Accuracy conditional formatting (red-white-green scale)
    last_row = len(pair_df) + 1
    for col_idx in [6, 7, 8]:
        col_letter = get_column_letter(col_idx)
        cell_range = f"{col_letter}2:{col_letter}{last_row}"
        ws.conditional_formatting.add(cell_range, ColorScaleRule(
            start_type="num", start_value=0,   start_color="FFC7CE",
            mid_type="num",   mid_value=50,    mid_color="FFFFFF",
            end_type="num",   end_value=100,   end_color="C6EFCE",
        ))

    # ── Bottom section: per-type summary ──
    if not backtest_df.empty:
        gap_row = len(pair_df) + 4
        ws.merge_cells(f"A{gap_row}:K{gap_row}")
        c = ws.cell(row=gap_row, column=1, value="  ACCURACY BY SIGNAL TYPE")
        c.font  = Font(name=FONT_NAME, bold=True, size=10, color="FFFFFF")
        c.fill  = hfill("2E4057")
        c.alignment = Alignment(vertical="center")
        ws.row_dimensions[gap_row].height = 20
        gap_row += 1

        type_headers = ["Signal Type", "Count", "Accuracy @5m %",
                        "Accuracy @15m %", "Accuracy @30m %", "Avg Profit %"]
        for ci, h in enumerate(type_headers, 1):
            c = ws.cell(row=gap_row, column=ci, value=h)
            c.font  = Font(name=FONT_NAME, bold=True, size=10, color="FFFFFF")
            c.fill  = hfill("4472C4")
            c.border = thin_border()
            c.alignment = Alignment(horizontal="center")
        gap_row += 1

        for stype, grp in backtest_df.groupby("signal_type"):
            def acc(col):
                valid = grp[col].dropna()
                return round((valid > 0).sum() / len(valid) * 100, 1) if len(valid) > 0 else None

            vals = [
                stype.replace("_", " ").title(),
                len(grp),
                acc("move_5m_pct"),
                acc("move_15m_pct"),
                acc("move_30m_pct"),
                round(grp["estimated_profit_pct"].mean(), 4),
            ]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=gap_row, column=ci, value=v)
                c.font   = body_font()
                c.border = thin_border()
                c.alignment = Alignment(horizontal="left" if ci == 1 else "center")
            gap_row += 1


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Export post-warm-up signal backtest from Jupiter monitor DB."
    )
    parser.add_argument(
        "--hours", type=int, default=8,
        help="How many hours back to search for the most recent session start. "
             "Increase this if you ran the export well after the session ended. "
             "(default: 8)"
    )
    args = parser.parse_args()

    print("\n╔══════════════════════════════════════════════════════╗")
    print("║      POST WARM-UP BACKTEST EXPORTER                 ║")
    print("╚══════════════════════════════════════════════════════╝\n")
    print(f"  Lookback window    : {args.hours}h "
          f"(use --hours N to extend if session is older)")

    conn = get_conn()

    # Derive session window using gap detection
    session_start, warmup_end, fresh_count = derive_session_window(conn, args.hours)
    now = datetime.now()

    print(f"  Session start      : {session_start.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  Warm-up duration   : {WARMUP_DURATION_S}s ({WARMUP_DURATION_S // 60} min, "
          f"{WARMUP_ITERATIONS} iterations × {ITERATION_INTERVAL_S}s)")
    print(f"  Warm-up ended at   : {warmup_end.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  Quotes in session  : {fresh_count}")

    if warmup_end > now:
        remaining = int((warmup_end - now).total_seconds())
        print(f"\n  ⚠️  Warm-up still in progress — {remaining}s remaining.")
        print(f"     Exporting what exists, but signal data will be sparse.\n")

    # Load data
    signals_df        = load_signals(conn, warmup_end)
    warmup_signals_df = load_warmup_signals(conn, session_start, warmup_end)
    prices_df         = load_prices(conn)

    print(f"\n  Warm-up signals (suppressed) : {len(warmup_signals_df)}")
    print(f"  Post warm-up signals         : {len(signals_df)}")

    if signals_df.empty:
        print("\n  ⚠️  No post-warm-up signals found.")
        print("     The session may be too new, or no signals have fired yet.")
        if not warmup_signals_df.empty:
            print(f"     ({len(warmup_signals_df)} warm-up signals were suppressed and excluded.)")

    # Build analysis frames
    backtest_df  = build_backtest(signals_df, prices_df) if not signals_df.empty else pd.DataFrame()
    pair_perf_df = build_pair_performance(backtest_df) if not backtest_df.empty else pd.DataFrame()

    # Build workbook
    wb = Workbook()
    write_summary_sheet(wb, session_start, warmup_end, signals_df,
                        warmup_signals_df, fresh_count)
    write_signals_sheet(wb, signals_df)
    write_backtest_sheet(wb, backtest_df)
    write_pair_performance_sheet(wb, pair_perf_df, backtest_df)

    # Save
    out_dir  = Path("price_history")
    out_dir.mkdir(exist_ok=True)
    filename = f"backtest_session_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    out_path = out_dir / filename

    wb.save(str(out_path))
    conn.close()

    print(f"\n  ✅ Exported: {out_path}")
    print(f"     Sheets: Session Summary | Signals | Backtest | Pair Performance")
    print(f"══════════════════════════════════════════════════════\n")


if __name__ == "__main__":
    main()
